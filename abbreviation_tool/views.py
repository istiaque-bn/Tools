from django.conf import settings
from django.contrib.auth.decorators import login_required, permission_required
from django.db import IntegrityError, models, transaction
from django.db.models import Count, IntegerField, OuterRef, Subquery
from accounts.decorators import admin_required
import csv
import io
from datetime import timedelta
import json
import logging
from pathlib import Path

from django.core.exceptions import ValidationError
from django.contrib import messages
from django.http import FileResponse, Http404, JsonResponse
from django.shortcuts import get_object_or_404, redirect, render
from django.utils import timezone

from .forms import AbbreviationEntryForm, DictionaryImportForm, DictionarySearchForm, DocumentUploadForm, FeedbackForm, PowerPointProcessForm, QuickProcessForm
from .models import AbbreviationAuditLog, AbbreviationEntry, AbbreviationProfile, DocumentProcessingSession
from .storage import cleanup_expired, expire_session, save_original
from .services.analysis import analyse_session
from .services.preview import build_preview
from .services.review import bulk_decide, decide, history_action
from .services.generation import generate_session, glossary_rows
from .storage import PROCESSED_NAME, session_directory


logger = logging.getLogger(__name__)


def _feature_enabled():
    return settings.DOCX_ABBREVIATION_TOOL_ENABLED


def _require_feature():
    if not _feature_enabled():
        raise Http404("SD Checker is disabled.")


@login_required
@permission_required("abbreviation_tool.access_abbreviation_tool", raise_exception=True)
def landing(request):
    _require_feature()
    cleanup_expired()
    if not request.user.has_perm("abbreviation_tool.process_document"):
        from django.core.exceptions import PermissionDenied
        raise PermissionDenied
    form = QuickProcessForm(request.POST or None, request.FILES or None)
    if request.method == "POST" and form.is_valid():
        active_sessions = DocumentProcessingSession.objects.filter(user=request.user, deleted_at__isnull=True, expires_at__gt=timezone.now()).count()
        if active_sessions >= settings.DOCX_ABBREVIATION_MAX_ACTIVE_SESSIONS:
            return render(request, "abbreviation_tool/landing.html", {"form": form, "powerpoint_form": PowerPointProcessForm(), "error": "Please wait for an existing document session to expire or cancel it."}, status=429)
        document = form.cleaned_data["docx_file"]
        profile = AbbreviationProfile.objects.filter(name="General", active=True).first() or AbbreviationProfile.objects.filter(active=True).first()
        session = DocumentProcessingSession.objects.create(
            user=request.user,
            original_filename=Path(document.name).name[:255],
            operation_type=form.cleaned_data["operation_type"],
            profile=profile,
            replacement_policy=DocumentProcessingSession.Policy.DEFINE_FIRST,
            processing_options={"include_tables": True, "include_headers_footers": True, "include_footnotes_endnotes": True, "glossary_mode": "none"},
            file_size=document.size,
            unsupported_element_count=len(form.inspection.unsupported_elements),
            expires_at=timezone.now() + timedelta(minutes=settings.DOCX_ABBREVIATION_SESSION_TTL_MINUTES),
        )
        try:
            save_original(session, document)
            analyse_session(session, policy=session.replacement_policy, include_tables=True)
            session.suggestions.filter(ambiguity_status="unambiguous").update(review_status="accepted")
            session.suggestions.filter(ambiguity_status="ambiguous").update(review_status="rejected")
            session.accepted_count = session.suggestions.filter(review_status="accepted").count()
            session.rejected_count = session.suggestions.filter(review_status="rejected").count()
            session.save(update_fields=("accepted_count", "rejected_count"))
            generate_session(session.id, request.user)
            path = session_directory(session.id) / PROCESSED_NAME
            filename = f"processed-{Path(session.original_filename).stem}.docx"
            response = FileResponse(path.open("rb"), as_attachment=True, filename=filename, content_type="application/vnd.openxmlformats-officedocument.wordprocessingml.document")
            response["Cache-Control"] = "no-store, private"
            response["X-Content-Type-Options"] = "nosniff"
            response._resource_closers.append(lambda: expire_session(session))
            return response
        except ValidationError as exc:
            expire_session(session, status=DocumentProcessingSession.Status.FAILED)
            return render(request, "abbreviation_tool/landing.html", {"form": form, "powerpoint_form": PowerPointProcessForm(), "error": exc.messages[0]}, status=400)
        except Exception:
            logger.exception("Quick DOCX processing failed", extra={"processing_session_id": str(session.id), "user_id": request.user.id})
            expire_session(session, status=DocumentProcessingSession.Status.FAILED)
            return render(request, "abbreviation_tool/landing.html", {"form": form, "powerpoint_form": PowerPointProcessForm(), "error": "The document could not be processed. Temporary files were deleted."}, status=400)
    return render(request, "abbreviation_tool/landing.html", {
        "form": form,
        "powerpoint_form": PowerPointProcessForm(),
        "entry_count": AbbreviationEntry.objects.filter(status=AbbreviationEntry.Status.ACTIVE).count(),
    })


@login_required
@permission_required("abbreviation_tool.process_document", raise_exception=True)
def powerpoint_convert(request):
    _require_feature()
    if request.method != "POST":
        return redirect("abbreviation_tool:landing")
    form = PowerPointProcessForm(request.POST, request.FILES)
    if not form.is_valid():
        return render(request, "abbreviation_tool/landing.html", {
            "form": QuickProcessForm(), "powerpoint_form": form,
            "entry_count": AbbreviationEntry.objects.filter(status=AbbreviationEntry.Status.ACTIVE).count(),
        }, status=400)
    profile = AbbreviationProfile.objects.filter(name="General", active=True).first()
    try:
        from .services.powerpoint import process_powerpoint
        output, count = process_powerpoint(form.cleaned_data["presentation_file"], form.cleaned_data["operation_type"], profile)
    except ValidationError as exc:
        form.add_error("presentation_file", exc.messages[0])
        return render(request, "abbreviation_tool/landing.html", {
            "form": QuickProcessForm(), "powerpoint_form": form,
            "entry_count": AbbreviationEntry.objects.filter(status=AbbreviationEntry.Status.ACTIVE).count(),
        }, status=400)
    source_name = Path(form.cleaned_data["presentation_file"].name).stem
    response = FileResponse(output, as_attachment=True, filename=f"processed-{source_name}.pptx", content_type="application/vnd.openxmlformats-officedocument.presentationml.presentation")
    response["X-Replacement-Count"] = str(count)
    response["X-Content-Type-Options"] = "nosniff"
    return response


@login_required
@permission_required("abbreviation_tool.process_document", raise_exception=True)
def text_convert(request):
    _require_feature()
    if request.method != "POST":
        return JsonResponse({"error": "Method not allowed."}, status=405)
    data = _payload(request)
    text = str(data.get("text", ""))
    operation = data.get("operation")
    if operation not in {DocumentProcessingSession.Operation.ABBREVIATE, DocumentProcessingSession.Operation.DEABBREVIATE}:
        return JsonResponse({"error": "Choose abbreviate or deabbreviate."}, status=400)
    if not text.strip() or len(text) > 50_000:
        return JsonResponse({"error": "Enter between 1 and 50,000 characters."}, status=400)
    from .services.matching import POLICY_ALL, candidates_for, find_matches
    from .services.ooxml import CharacterLocation, TextContainer
    profile = AbbreviationProfile.objects.filter(name="General", active=True).first()
    container = TextContainer("text", "text:p0", "paragraph", text, [CharacterLocation(0, index, b"") for index in range(len(text))])
    matches = find_matches(container, candidates_for(profile, operation), operation, POLICY_ALL)
    if operation == DocumentProcessingSession.Operation.DEABBREVIATE:
        matches = [match for match in matches if match.ambiguity == "unambiguous"]
    output = text
    for match in sorted(matches, key=lambda item: item.start, reverse=True):
        output = output[:match.start] + match.proposed + output[match.end:]
    return JsonResponse({"result": output, "replacement_count": len(matches)})


@login_required
@permission_required("abbreviation_tool.search_dictionary", raise_exception=True)
def dictionary(request):
    _require_feature()
    form = DictionarySearchForm(request.GET)
    entries = AbbreviationEntry.objects.select_related("category").prefetch_related("profiles")
    if form.is_valid():
        query = form.cleaned_data.get("q")
        if query:
            entries = entries.filter(models.Q(abbreviation__icontains=query) | models.Q(full_form__icontains=query))
        if form.cleaned_data.get("service"):
            entries = entries.filter(service__iexact=form.cleaned_data["service"])
        if form.cleaned_data.get("status"):
            entries = entries.filter(status=form.cleaned_data["status"])
        if form.cleaned_data.get("ambiguous") is not None:
            entries = entries.filter(is_ambiguous=form.cleaned_data["ambiguous"])
    return render(request, "abbreviation_tool/dictionary.html", {"form": form, "entries": entries[:250]})


@admin_required
def manage_dictionary(request, entry_id=None):
    _require_feature()
    entry = get_object_or_404(AbbreviationEntry, pk=entry_id) if entry_id else None
    form = AbbreviationEntryForm(request.POST or None, instance=entry, prefix="entry")
    import_form = DictionaryImportForm(request.POST or None, request.FILES or None, prefix="import")
    if request.method == "POST" and "save_entry" in request.POST and form.is_valid():
        before = None if entry is None else {"abbreviation": entry.abbreviation, "full_form": entry.full_form}
        saved = form.save(commit=False)
        saved.created_by = saved.created_by or request.user
        saved.updated_by = request.user
        try:
            saved.save()
        except IntegrityError:
            form.add_error(None, "This abbreviation and full-form pair already exists.")
        else:
            general_profile = AbbreviationProfile.objects.filter(name="General", active=True).first()
            if general_profile:
                saved.profiles.add(general_profile)
            AbbreviationAuditLog.objects.create(abbreviation_entry=saved, action="updated" if entry else "created", previous_value=before, new_value={"abbreviation": saved.abbreviation, "full_form": saved.full_form}, user=request.user)
            messages.success(request, "Abbreviation updated." if entry else "Abbreviation added.")
            return redirect("abbreviation_tool:manage_dictionary")
    if request.method == "POST" and "import_entries" in request.POST and import_form.is_valid():
        result = _import_dictionary_file(import_form.cleaned_data["file"], request.user)
        for error in result["errors"][:20]:
            messages.error(request, error)
        messages.success(request, f"Import complete: {result['created']} added, {result['updated']} updated, {result['skipped']} skipped.")
        return redirect("abbreviation_tool:manage_dictionary")
    query = " ".join(request.GET.get("q", "").split())[:200]
    total_entries = AbbreviationEntry.objects.count()
    meaning_counts = (
        AbbreviationEntry.objects.filter(normalized_abbreviation=OuterRef("normalized_abbreviation"))
        .values("normalized_abbreviation").annotate(total=Count("normalized_full_form", distinct=True)).values("total")
    )
    managed_entries = AbbreviationEntry.objects.annotate(meaning_count=Subquery(meaning_counts, output_field=IntegerField()))
    all_entries = managed_entries.order_by("full_form", "abbreviation")[:500]
    search_results = AbbreviationEntry.objects.none()
    if query:
        search_results = managed_entries.filter(models.Q(abbreviation__icontains=query) | models.Q(full_form__icontains=query)).order_by("full_form", "abbreviation")[:500]
    return render(request, "abbreviation_tool/manage_dictionary.html", {"form": form, "import_form": import_form, "entries": all_entries, "search_results": search_results, "editing": entry, "query": query, "total_entries": total_entries})


@admin_required
def export_dictionary_xlsx(request):
    _require_feature()
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill

    workbook = Workbook(write_only=False)
    worksheet = workbook.active
    worksheet.title = "Abbreviations"
    worksheet.append(("abbreviation", "full_form"))
    for abbreviation, full_form in AbbreviationEntry.objects.order_by("abbreviation", "full_form").values_list("abbreviation", "full_form").iterator():
        worksheet.append((abbreviation, full_form))

    header_fill = PatternFill("solid", fgColor="173564")
    for cell in worksheet[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = header_fill
    worksheet.freeze_panes = "A2"
    worksheet.auto_filter.ref = worksheet.dimensions
    worksheet.column_dimensions["A"].width = 24
    worksheet.column_dimensions["B"].width = 72

    output = io.BytesIO()
    workbook.save(output)
    output.seek(0)
    response = FileResponse(
        output,
        as_attachment=True,
        filename=f"abbreviations-{timezone.localdate().isoformat()}.xlsx",
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    response["X-Content-Type-Options"] = "nosniff"
    response["Cache-Control"] = "no-store, private"
    return response


@admin_required
def delete_dictionary_entry(request, entry_id):
    _require_feature()
    if request.method != "POST":
        return redirect("abbreviation_tool:manage_dictionary")
    entry = get_object_or_404(AbbreviationEntry, pk=entry_id)
    label = f"{entry.abbreviation} — {entry.full_form}"
    AbbreviationAuditLog.objects.create(
        abbreviation_entry=entry,
        action="deleted",
        previous_value={"abbreviation": entry.abbreviation, "full_form": entry.full_form, "status": entry.status},
        user=request.user,
    )
    entry.delete()
    messages.success(request, f'Abbreviation "{label}" was deleted successfully.')
    return redirect("abbreviation_tool:manage_dictionary")


def _import_dictionary_file(upload, user):
    result = {"created": 0, "updated": 0, "skipped": 0, "errors": []}
    if upload.name.lower().endswith(".xlsx"):
        from openpyxl import load_workbook
        try:
            worksheet = load_workbook(upload, read_only=True, data_only=True).active
            values = worksheet.iter_rows(values_only=True)
            first_row = next(values, None)
            if not first_row:
                result["errors"].append("The spreadsheet is empty.")
                return result
            fieldnames = [str(value).strip() if value is not None else "" for value in first_row]
            reader = ({fieldnames[index]: value for index, value in enumerate(row) if index < len(fieldnames)} for row in values)
        except Exception:
            result["errors"].append("The XLSX file is invalid or cannot be read.")
            return result
    else:
        try:
            reader = csv.DictReader(io.StringIO(upload.read().decode("utf-8-sig"), newline=""))
            fieldnames = reader.fieldnames or []
        except UnicodeDecodeError:
            result["errors"].append("The CSV file must use UTF-8 encoding.")
            return result
    headers = {str(name).strip().casefold(): name for name in fieldnames}
    short_key = headers.get("abbreviation") or headers.get("short form") or headers.get("short_form")
    full_key = headers.get("full_form") or headers.get("full form")
    if not short_key or not full_key:
        result["errors"].append("CSV headers must include abbreviation and full_form.")
        return result
    for line, row in enumerate(reader, 2):
        abbreviation = " ".join(str(row.get(short_key) or "").split())
        full_form = " ".join(str(row.get(full_key) or "").split())
        if " / " in full_form:
            result["skipped"] += 1
            result["errors"].append(f"Row {line}: use a separate row for each full form instead of '/'.")
            continue
        if not abbreviation or not full_form or len(abbreviation) > 100 or len(full_form) > 500:
            result["skipped"] += 1
            result["errors"].append(f"Row {line}: invalid or missing value.")
            continue
        existing = AbbreviationEntry.objects.filter(normalized_full_form=AbbreviationEntry.normalize(full_form)).order_by("pk").first()
        if existing and existing.normalized_abbreviation == AbbreviationEntry.normalize(abbreviation):
            result["skipped"] += 1
            continue
        before = {"abbreviation": existing.abbreviation, "full_form": existing.full_form} if existing else None
        entry = existing or AbbreviationEntry(created_by=user)
        entry.abbreviation, entry.full_form, entry.updated_by = abbreviation, full_form, user
        entry.source_name = entry.source_name or "CSV import"
        try:
            with transaction.atomic():
                entry.save()
        except IntegrityError:
            result["skipped"] += 1
            result["errors"].append(f"Row {line}: conflicts with an existing pair.")
            continue
        AbbreviationAuditLog.objects.create(abbreviation_entry=entry, action="import_updated" if existing else "import_created", previous_value=before, new_value={"abbreviation": abbreviation, "full_form": full_form}, user=user)
        general_profile = AbbreviationProfile.objects.filter(name="General", active=True).first()
        if general_profile:
            entry.profiles.add(general_profile)
        result["updated" if existing else "created"] += 1
    return result


@login_required
def feedback(request):
    form = FeedbackForm(request.POST or None, initial={"name": request.user.get_full_name() or request.user.username, "email": request.user.email})
    if request.method == "POST" and form.is_valid():
        item = form.save(commit=False)
        item.user = request.user
        item.save()
        messages.success(request, "Thank you. Your feedback was sent to the administrator.")
        return redirect("abbreviation_tool:feedback")
    return render(request, "abbreviation_tool/feedback.html", {"form": form})


@login_required
@permission_required("abbreviation_tool.process_document", raise_exception=True)
def upload(request):
    _require_feature()
    active_sessions = DocumentProcessingSession.objects.filter(user=request.user, deleted_at__isnull=True, expires_at__gt=timezone.now()).count()
    if request.method == "POST" and active_sessions >= settings.DOCX_ABBREVIATION_MAX_ACTIVE_SESSIONS:
        return render(request, "abbreviation_tool/upload.html", {"form": DocumentUploadForm(), "limit_error": "Finish or cancel an existing document session before starting another."}, status=429)
    form = DocumentUploadForm(request.POST or None, request.FILES or None)
    if request.method == "POST" and form.is_valid():
        document = form.cleaned_data["pdf_file"]
        session = DocumentProcessingSession.objects.create(
            user=request.user,
            original_filename=Path(document.name).name[:255],
            operation_type=form.cleaned_data["operation_type"],
            profile=form.cleaned_data["profile"],
            replacement_policy=form.cleaned_data["replacement_policy"],
            processing_options={
                "include_tables": form.cleaned_data["include_tables"],
                "include_headers_footers": form.cleaned_data["include_headers_footers"],
                "include_footnotes_endnotes": form.cleaned_data["include_footnotes_endnotes"],
                "case_sensitive": form.cleaned_data["case_sensitive"],
                "high_confidence_only": form.cleaned_data["high_confidence_only"],
                "glossary_mode": form.cleaned_data["glossary_mode"] or "none",
                "glossary_bookmark": form.cleaned_data["glossary_bookmark"],
            },
            file_size=document.size,
            unsupported_element_count=len(form.inspection.unsupported_elements),
            expires_at=timezone.now() + timedelta(minutes=settings.DOCX_ABBREVIATION_SESSION_TTL_MINUTES),
        )
        try:
            save_original(session, document)
        except Exception:
            session.delete()
            raise
        return redirect("abbreviation_tool:session", session_id=session.id)
    return render(request, "abbreviation_tool/upload.html", {"form": form})


@login_required
@permission_required("abbreviation_tool.process_document", raise_exception=True)
def session_detail(request, session_id):
    _require_feature()
    session = get_object_or_404(DocumentProcessingSession, id=session_id, user=request.user, deleted_at__isnull=True)
    if session.expires_at <= timezone.now():
        expire_session(session)
        raise Http404("This processing session has expired.")
    return render(request, "abbreviation_tool/session.html", {"processing_session": session})


@login_required
@permission_required("abbreviation_tool.process_document", raise_exception=True)
def cancel(request, session_id):
    _require_feature()
    if request.method != "POST":
        raise Http404
    session = get_object_or_404(DocumentProcessingSession, id=session_id, user=request.user, deleted_at__isnull=True)
    expire_session(session)
    return redirect("abbreviation_tool:landing")


@login_required
@permission_required("abbreviation_tool.process_document", raise_exception=True)
def analyse(request, session_id):
    _require_feature()
    if request.method != "POST":
        raise Http404
    session = get_object_or_404(DocumentProcessingSession, id=session_id, user=request.user, deleted_at__isnull=True)
    if session.expires_at <= timezone.now():
        expire_session(session)
        raise Http404("This processing session has expired.")
    try:
        analyse_session(session, policy=session.replacement_policy, include_tables=session.processing_options.get("include_tables", True))
        return redirect("abbreviation_tool:review", session_id=session.id)
    except ValidationError as exc:
        expire_session(session, status=DocumentProcessingSession.Status.FAILED)
        messages.error(request, exc.messages[0])
        return redirect("abbreviation_tool:landing")
    except Exception:
        logger.exception("DOCX analysis failed", extra={"processing_session_id": str(session.id), "user_id": request.user.id})
        expire_session(session, status=DocumentProcessingSession.Status.FAILED)
        messages.error(request, "Document analysis failed and its temporary files were deleted.")
        return redirect("abbreviation_tool:landing")


def _owned_review_session(user, session_id):
    session = get_object_or_404(DocumentProcessingSession.objects.prefetch_related("suggestions__abbreviation_entry__category"), id=session_id, user=user, deleted_at__isnull=True)
    if session.expires_at <= timezone.now():
        expire_session(session)
        raise Http404("This processing session has expired.")
    if session.status != DocumentProcessingSession.Status.REVIEW:
        raise Http404("This session is not ready for review.")
    return session


@login_required
@permission_required("abbreviation_tool.process_document", raise_exception=True)
def review(request, session_id):
    _require_feature()
    session = _owned_review_session(request.user, session_id)
    suggestions = session.suggestions.select_related("abbreviation_entry__category", "selected_meaning").order_by("container_identifier", "start_offset")
    ambiguity_options = {}
    for suggestion in suggestions.filter(ambiguity_status="ambiguous"):
        ambiguity_options[str(suggestion.id)] = list(AbbreviationEntry.objects.filter(
            normalized_abbreviation=suggestion.abbreviation_entry.normalized_abbreviation,
            status=AbbreviationEntry.Status.ACTIVE,
        ).values("id", "full_form"))
    return render(request, "abbreviation_tool/review.html", {
        "processing_session": session,
        "suggestions": suggestions,
        "preview": build_preview(session),
        "ambiguity_options": ambiguity_options,
        "categories": suggestions.exclude(abbreviation_entry__category__isnull=True).values_list("abbreviation_entry__category__name", flat=True).distinct(),
        "glossary": glossary_rows(session),
    })


def _payload(request):
    try:
        return json.loads(request.body or b"{}")
    except json.JSONDecodeError as exc:
        raise ValidationError("Invalid JSON request.") from exc


@login_required
@permission_required("abbreviation_tool.process_document", raise_exception=True)
def suggestion_api(request, session_id, suggestion_id):
    _require_feature()
    if request.method not in {"PATCH", "POST"}:
        return JsonResponse({"error": "Method not allowed."}, status=405)
    session = _owned_review_session(request.user, session_id)
    data = _payload(request)
    try:
        suggestion = decide(session_id, request.user, suggestion_id, data.get("action"), data.get("replacement", ""), data.get("selected_meaning_id"))
        return JsonResponse({"ok": True, "status": suggestion.review_status, "replacement": suggestion.user_modified_text or suggestion.proposed_text})
    except (ValidationError, AbbreviationEntry.DoesNotExist) as exc:
        message = exc.messages[0] if isinstance(exc, ValidationError) else "The selected meaning is invalid."
        return JsonResponse({"error": message}, status=400)


@login_required
@permission_required("abbreviation_tool.process_document", raise_exception=True)
def bulk_review_api(request, session_id):
    _require_feature()
    if request.method != "POST":
        return JsonResponse({"error": "Method not allowed."}, status=405)
    _owned_review_session(request.user, session_id)
    data = _payload(request)
    try:
        bulk_decide(session_id, request.user, data.get("action"), bool(data.get("high_confidence")), data.get("suggestion_ids"))
        return JsonResponse({"ok": True})
    except ValidationError as exc:
        return JsonResponse({"error": exc.messages[0]}, status=400)


@login_required
@permission_required("abbreviation_tool.process_document", raise_exception=True)
def history_api(request, session_id):
    _require_feature()
    if request.method != "POST":
        return JsonResponse({"error": "Method not allowed."}, status=405)
    _owned_review_session(request.user, session_id)
    try:
        history_action(session_id, request.user, _payload(request).get("direction"))
        return JsonResponse({"ok": True})
    except ValidationError as exc:
        return JsonResponse({"error": exc.messages[0]}, status=400)


@login_required
@permission_required("abbreviation_tool.process_document", raise_exception=True)
def generate(request, session_id):
    _require_feature()
    if request.method != "POST":
        raise Http404
    session = _owned_review_session(request.user, session_id)
    try:
        generate_session(session_id, request.user)
        return redirect("abbreviation_tool:summary", session_id=session_id)
    except ValidationError as exc:
        session.refresh_from_db()
        if session.status == DocumentProcessingSession.Status.FAILED:
            expire_session(session, status=DocumentProcessingSession.Status.FAILED)
            messages.error(request, f"{exc.messages[0]} Temporary files were deleted.")
            return redirect("abbreviation_tool:landing")
        messages.error(request, exc.messages[0])
        return redirect("abbreviation_tool:review", session_id=session_id)
    except Exception:
        logger.exception("DOCX generation failed", extra={"processing_session_id": str(session_id), "user_id": request.user.id})
        session = get_object_or_404(DocumentProcessingSession, id=session_id, user=request.user)
        expire_session(session, status=DocumentProcessingSession.Status.FAILED)
        messages.error(request, "DOCX generation failed and its temporary files were deleted.")
        return redirect("abbreviation_tool:landing")


@login_required
@permission_required("abbreviation_tool.process_document", raise_exception=True)
def summary(request, session_id):
    _require_feature()
    session = get_object_or_404(DocumentProcessingSession, id=session_id, user=request.user, status=DocumentProcessingSession.Status.COMPLETE, deleted_at__isnull=True)
    return render(request, "abbreviation_tool/summary.html", {"processing_session": session})


@login_required
@permission_required("abbreviation_tool.process_document", raise_exception=True)
def download(request, session_id):
    _require_feature()
    session = get_object_or_404(DocumentProcessingSession, id=session_id, user=request.user, status=DocumentProcessingSession.Status.COMPLETE, deleted_at__isnull=True)
    path = session_directory(session.id) / PROCESSED_NAME
    if not path.is_file():
        raise Http404("The generated document is no longer available.")
    filename = f"processed-{Path(session.original_filename).stem}.docx"
    response = FileResponse(path.open("rb"), as_attachment=True, filename=filename, content_type="application/vnd.openxmlformats-officedocument.wordprocessingml.document")
    response["Cache-Control"] = "no-store, private"
    response["X-Content-Type-Options"] = "nosniff"
    response._resource_closers.append(lambda: expire_session(session))
    return response


@login_required
@permission_required("abbreviation_tool.process_document", raise_exception=True)
def glossary_download(request, session_id):
    _require_feature()
    session = get_object_or_404(DocumentProcessingSession, id=session_id, user=request.user, deleted_at__isnull=True)
    if session.processing_options.get("glossary_mode") not in {"preview", "separate"}:
        raise Http404("A separate glossary was not selected for this session.")
    rows = glossary_rows(session)
    if not rows:
        return JsonResponse({"error": "No glossary entries are available."}, status=400)
    from django.http import HttpResponse
    content = "Abbreviation\tFull Form\n" + "\n".join(f"{abbreviation}\t{full_form}" for abbreviation, full_form in rows)
    response = HttpResponse(content, content_type="text/tab-separated-values; charset=utf-8")
    response["Content-Disposition"] = 'attachment; filename="abbreviation-glossary.tsv"'
    response["Cache-Control"] = "no-store, private"
    response["X-Content-Type-Options"] = "nosniff"
    return response
